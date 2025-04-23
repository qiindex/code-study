import pandas as pd
import json
import os
from openpyxl import load_workbook
from datetime import datetime
import numpy as np

# 定义字段映射关系
FIELD_MAPPING = {
    "应用": "application_type",
    "场景": "business_type",
    "功能": "functional_category",
    "接口": "url",
    "错误码": "error_code",
    "错误信息": "error_msg",
    "新错误码": "new_error_code",
    "FE、Client是否使用": "if_fe_used",
    "错误类型": "error_category",
    "紧急度": "urgency_degree_type",
    "原因列表": "error_reason_list",
    "告警策略": "error_alarm_strategy",
    "告警SOP": "error_alarm_sop",
    "Todo": "error_extra_info"
}

# 需要前向填充的字段列表
FILL_FIELDS = [
    'application_type',
    'business_type',
    'functional_category',
    'url',
    'error_code'
]

def process_merged_cells(excel_file_path):
    """
    处理Excel中的合并单元格，将第一行合并单元格的内容复制到该列后续的合并单元格中

    Args:
        excel_file_path (str): Excel文件的路径

    Returns:
        pandas.DataFrame: 处理后的DataFrame
    """
    # 使用openpyxl加载工作簿
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # 获取所有合并单元格
    merged_cells = ws.merged_cells.ranges

    # 创建一个字典来存储合并单元格的信息
    merge_info = {}

    # 处理每个合并单元格
    for merged_range in merged_cells:
        # 获取合并单元格的起始行和列
        start_row = merged_range.min_row
        start_col = merged_range.min_col
        end_row = merged_range.max_row
        end_col = merged_range.max_col

        # 获取合并单元格的值
        value = ws.cell(row=start_row, column=start_col).value

        # 如果是第一行的合并单元格，记录信息
        if start_row == 1:
            merge_info[start_col] = {
                'value': value,
                'end_col': end_col
            }

    # 将工作簿转换为DataFrame
    df = pd.read_excel(excel_file_path)

    # 处理合并单元格
    for start_col, info in merge_info.items():
        value = info['value']
        end_col = info['end_col']

        # 将值复制到该列的所有单元格
        for col in range(start_col, end_col + 1):
            df.iloc[:, col-1] = value

    return df

def excel_to_json(excel_file_path, output_json_path=None):
    """
    读取Excel文件并转换为JSON格式

    Args:
        excel_file_path (str): Excel文件的路径
        output_json_path (str, optional): 输出JSON文件的路径。如果为None，则使用Excel文件名

    Returns:
        dict: 转换后的JSON数据
    """
    # 处理合并单元格
    df = process_merged_cells(excel_file_path)

    # 重命名列名
    df = df.rename(columns=FIELD_MAPPING)

    # 对指定字段进行前向填充
    for field in FILL_FIELDS:
        df[field] = df[field].fillna(method='ffill')

    # 只对error_code字段进行浮点数转整数处理
    if 'error_code' in df.columns and df['error_code'].dtype == 'float64':
        df['error_code'] = df['error_code'].apply(lambda x: int(x) if pd.notnull(x) else x)

    # 过滤掉error_msg为空或null的行
    df = df.dropna(subset=['error_msg'])

    # 将DataFrame转换为字典，并将NaN替换为None
    data = df.replace({np.nan: None}).to_dict(orient='records')

    # 将数据包装在list键中
    wrapped_data = {"list": data}

    # 如果没有指定输出路径，则使用Excel文件名加上时间戳
    if output_json_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.splitext(os.path.basename(excel_file_path))[0]
        output_json_path = f"{base_name}_{timestamp}.json"

    # 将数据写入JSON文件
    with open(output_json_path, 'w', encoding='utf-8') as f:
        json.dump(wrapped_data, f, ensure_ascii=False, indent=4)

    return wrapped_data

if __name__ == "__main__":
    # 示例用法
    excel_file = "/Users/yuanqiang.qi/VsCode/2025-github/new-python-cursor/Account告警准确性治理.xlsx"  # 使用完整路径
    try:
        result = excel_to_json(excel_file)
        print(f"Excel数据已成功转换为JSON格式,并保存到 {os.path.basename(excel_file)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
    except Exception as e:
        print(f"转换过程中出现错误: {str(e)}")